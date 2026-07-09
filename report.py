import io
from datetime import datetime

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

# ==========================================================
# DataFrame Resumen
# ==========================================================

def build_summary(result,
                  customer,
                  project,
                  environment,
                  architecture,
                  storage):

    return pd.DataFrame({

        "Parámetro":[

            "Cliente",
            "Proyecto",
            "Ambiente",
            "Arquitectura",
            "Storage",

            "CPU Requerida",

            "RAM Requerida",

            "Storage Requerido",

            "Masters",

            "Workers",

            "Infra Nodes",

            "Total Nodes",

            "Suscripciones",

            "Assessment Score"

        ],

        "Valor":[

            customer,

            project,

            environment,

            architecture,

            storage,

            result["cpu"],

            result["ram"],

            result["storage"],

            result["masters"],

            result["workers"],

            result["infra"],

            result["total_nodes"],

            result["subscriptions"],

            str(result["score"])+" %"

        ]

    })


# ==========================================================
# EXCEL
# ==========================================================

def export_excel(df):

    wb = Workbook()

    ws = wb.active

    ws.title = "Sizing"

    blue_fill = PatternFill(
        start_color="C00000",
        end_color="C00000",
        fill_type="solid"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    for col, name in enumerate(df.columns, start=1):

        cell = ws.cell(
            row=1,
            column=col
        )

        cell.value = name

        cell.fill = blue_fill

        cell.font = white_font

    for row in df.itertuples(index=False):

        ws.append(row)

    stream = io.BytesIO()

    wb.save(stream)

    stream.seek(0)

    return stream


# ==========================================================
# PDF
# ==========================================================

def export_pdf(df):

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()

    elements = []

    title = Paragraph(
        "<b>OpenShift Virtualization Assessment Report</b>",
        styles["Title"]
    )

    elements.append(title)

    elements.append(Spacer(1,20))

    date = Paragraph(
        datetime.now().strftime("%d/%m/%Y"),
        styles["Normal"]
    )

    elements.append(date)

    elements.append(Spacer(1,20))

    data = [df.columns.tolist()] + df.values.tolist()

    table = Table(data)

    table.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.red),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("GRID",(0,0),(-1,-1),1,colors.grey),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

            ("BOTTOMPADDING",(0,0),(-1,0),10),

            ("BACKGROUND",(0,1),(-1,-1),colors.beige)

        ])

    )

    elements.append(table)

    elements.append(Spacer(1,30))

    elements.append(

        Paragraph(

            "<b>Executive Summary</b>",

            styles["Heading2"]

        )

    )

    elements.append(

        Paragraph(

            """
            El assessment realizado estima la infraestructura
            necesaria para implementar OpenShift Virtualization,
            considerando capacidad de cómputo, almacenamiento,
            crecimiento proyectado y disponibilidad.

            La arquitectura recomendada contempla un clúster de
            alta disponibilidad basado en Kubernetes con tres
            nodos de control (Masters) y los nodos Worker
            requeridos para alojar las máquinas virtuales.

            """,

            styles["BodyText"]

        )

    )

    doc.build(elements)

    buffer.seek(0)

    return buffer


# ==========================================================
# BOTONES STREAMLIT
# ==========================================================

def show_download_buttons(df):

    csv = df.to_csv(index=False).encode("utf-8")

    st.download_button(

        "📄 Descargar CSV",

        csv,

        "OpenShiftSizing.csv",

        "text/csv"

    )

    excel = export_excel(df)

    st.download_button(

        "📗 Descargar Excel",

        excel,

        "OpenShiftSizing.xlsx",

        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    pdf = export_pdf(df)

    st.download_button(

        "📕 Descargar PDF",

        pdf,

        "OpenShiftSizing.pdf",

        "application/pdf"

    )
